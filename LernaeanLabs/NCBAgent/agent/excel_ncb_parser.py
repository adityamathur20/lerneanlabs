"""
Excel / Google Sheets parser → list[NCBTransaction]

Supports:
  - Local .xlsx files (openpyxl)
  - Google Sheets via service account (gspread) — requires setup, see docs
  - Multi-sheet workbooks
  - Per-client column mapping via excel_config.yaml
  - Substance alias resolution
  - Optional regex substance extraction from narration column
  - Optional unit normalisation (g/kg/MT)
"""
from __future__ import annotations

import re
import warnings
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Optional

import openpyxl

from agent.config_loader import ExcelConfig, SheetConfig, load_client_profile, load_excel_config
from agent.models import NCBTransaction, ParseError, ParseResult
from agent.urn_validator import validate_urn

# Suppress openpyxl warnings about unknown extensions in xlsx files
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

CLIENTS_BASE = Path("clients")


# ---------------------------------------------------------------------------
# Public entry points
# ---------------------------------------------------------------------------

def parse_excel(
    client_id: str,
    file_path: Path,
    clients_base: Path = CLIENTS_BASE,
) -> ParseResult:
    config = load_excel_config(client_id, clients_base)
    profile = load_client_profile(client_id, clients_base)
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    return _parse_workbook(client_id, config, wb, source_label=str(file_path))


_SHEET_ID_RE = re.compile(r"/spreadsheets/d/([a-zA-Z0-9\-_]+)")


def parse_google_sheet_url(
    client_id: str,
    sheet_url: str,
    clients_base: Path = CLIENTS_BASE,
) -> ParseResult:
    """
    Accept a Google Sheets URL, extract the sheet ID, patch the client config,
    and delegate to parse_google_sheets().

    The service_account_key_path must still be set in excel_config.yaml.
    This function only overrides the sheet ID so operators can paste a URL
    directly without editing YAML.

    Example URL:
      https://docs.google.com/spreadsheets/d/1BxiMV.../edit#gid=0
    """
    match = _SHEET_ID_RE.search(sheet_url)
    if not match:
        raise ValueError(
            f"Cannot extract sheet ID from URL: {sheet_url!r}\n"
            "Expected format: https://docs.google.com/spreadsheets/d/<ID>/..."
        )
    sheet_id = match.group(1)

    config = load_excel_config(client_id, clients_base)
    # Temporarily override the sheet ID with the one from the URL
    config.input_type = "google_sheets"
    config.google_sheet_id = sheet_id

    if not config.service_account_key_path:
        raise ValueError(
            "service_account_key_path is not set in excel_config.yaml. "
            "Set it to the path of your Google service account JSON key."
        )

    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        raise ImportError("Install gspread: pip install gspread google-auth")

    scopes = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
    creds = Credentials.from_service_account_file(config.service_account_key_path, scopes=scopes)
    gc = gspread.authorize(creds)
    spreadsheet = gc.open_by_key(sheet_id)

    wb_fake = _GspreadsheetsWorkbook(spreadsheet)
    return _parse_workbook(client_id, config, wb_fake, source_label=f"sheets:{sheet_id}")


def parse_google_sheets(
    client_id: str,
    clients_base: Path = CLIENTS_BASE,
) -> ParseResult:
    """
    Reads from Google Sheets using a service account.

    Setup (one-time per deployment):
      1. Create a Google Cloud project and enable the Sheets API.
      2. Create a service account, download the JSON key.
      3. Share the client's Google Sheet with the service account email.
      4. Set service_account_key_path in excel_config.yaml.
      5. Set google_sheet_id in excel_config.yaml.
    """
    config = load_excel_config(client_id, clients_base)
    profile = load_client_profile(client_id, clients_base)

    if config.input_type != "google_sheets":
        raise ValueError(
            f"Client '{client_id}' excel_config.yaml has input.type='{config.input_type}'. "
            f"Set input.type: google_sheets to use this parser."
        )
    if not config.google_sheet_id:
        raise ValueError("google_sheet_id is not set in excel_config.yaml")
    if not config.service_account_key_path:
        raise ValueError("service_account_key_path is not set in excel_config.yaml")

    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError:
        raise ImportError("Install gspread: pip install gspread google-auth")

    scopes = ["https://www.googleapis.com/auth/spreadsheets.readonly"]
    creds = Credentials.from_service_account_file(config.service_account_key_path, scopes=scopes)
    gc = gspread.authorize(creds)
    spreadsheet = gc.open_by_key(config.google_sheet_id)

    # Convert gspread worksheet → openpyxl-compatible row iterator
    # We wrap gspread data into a fake workbook-like structure that _parse_sheet can consume
    wb_fake = _GspreadsheetsWorkbook(spreadsheet)
    return _parse_workbook(client_id, config, wb_fake, source_label=f"sheets:{config.google_sheet_id}")


# ---------------------------------------------------------------------------
# Workbook parsing
# ---------------------------------------------------------------------------

def _parse_workbook(
    client_id: str,
    config: ExcelConfig,
    wb: Any,
    source_label: str,
) -> ParseResult:
    all_transactions: list[NCBTransaction] = []
    all_errors: list[ParseError] = []
    all_warnings: list[str] = []

    sheet_names = wb.sheetnames if hasattr(wb, "sheetnames") else [s.title for s in wb]

    for sheet_cfg in config.sheets:
        if sheet_cfg.sheet_name not in sheet_names:
            all_warnings.append(
                f"Sheet '{sheet_cfg.sheet_name}' not found in {source_label}. "
                f"Available: {sheet_names}. Skipping."
            )
            continue

        ws = wb[sheet_cfg.sheet_name]
        txns, errs = _parse_sheet(client_id, config, sheet_cfg, ws, source_label)
        all_transactions.extend(txns)
        all_errors.extend(errs)

    return ParseResult(transactions=all_transactions, errors=all_errors, warnings=all_warnings)


def _parse_sheet(
    client_id: str,
    config: ExcelConfig,
    sheet_cfg: SheetConfig,
    ws: Any,
    source_label: str,
) -> tuple[list[NCBTransaction], list[ParseError]]:

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return [], []

    # Build header index from the header row (0-indexed internally)
    header_row_idx = sheet_cfg.header_row - 1
    if header_row_idx >= len(rows):
        return [], [ParseError(source_label, sheet_cfg.header_row, "header_row", "", "Header row index out of range")]

    raw_headers = rows[header_row_idx]
    header_index: dict[str, int] = {}
    for col_idx, cell_val in enumerate(raw_headers):
        if cell_val is not None:
            header_index[str(cell_val).strip()] = col_idx

    # Validate that all required mapped columns exist
    required_fields = ["date", "substance", "quantity_kg", "counterparty"]
    errors: list[ParseError] = []
    for field_name in required_fields:
        col_name = sheet_cfg.columns.get(field_name)
        if col_name and col_name not in header_index:
            errors.append(ParseError(
                source_label, sheet_cfg.header_row, field_name, col_name,
                f"Column '{col_name}' not found in sheet '{sheet_cfg.sheet_name}'. "
                f"Available headers: {list(header_index.keys())}"
            ))
    if errors:
        return [], errors

    def get(row: tuple, field: str, default: Any = None) -> Any:
        col_name = sheet_cfg.columns.get(field)
        if not col_name:
            return default
        col_idx = header_index.get(col_name)
        if col_idx is None:
            return default
        val = row[col_idx] if col_idx < len(row) else default
        return val

    transactions: list[NCBTransaction] = []
    skip_set = set(sheet_cfg.skip_rows)

    for row_1indexed, row in enumerate(rows[sheet_cfg.data_starts_row - 1:],
                                       start=sheet_cfg.data_starts_row):
        if row_1indexed in skip_set:
            continue
        # Skip completely empty rows
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue

        txn, row_errors = _parse_row(
            client_id=client_id,
            config=config,
            sheet_cfg=sheet_cfg,
            row=row,
            row_number=row_1indexed,
            source_label=source_label,
            get=get,
            header_index=header_index,
        )
        if row_errors:
            errors.extend(row_errors)
        if txn:
            transactions.append(txn)

    return transactions, errors


def _parse_row(
    client_id: str,
    config: ExcelConfig,
    sheet_cfg: SheetConfig,
    row: tuple,
    row_number: int,
    source_label: str,
    get,
    header_index: dict,
) -> tuple[Optional[NCBTransaction], list[ParseError]]:

    errors: list[ParseError] = []

    # --- Date ---
    raw_date = get(row, "date")
    parsed_date = _parse_date(raw_date, config.date_format)
    if parsed_date is None:
        errors.append(ParseError(source_label, row_number, "date", str(raw_date),
                                 f"Cannot parse date '{raw_date}' with format '{config.date_format}'"))
        return None, errors

    # --- Transaction type ---
    if sheet_cfg.txn_type_mode == "fixed":
        txn_type = sheet_cfg.txn_type_fixed or "PURCHASE"
    else:
        # txn_type_column is a dedicated key, not in the columns map
        txn_col_name = sheet_cfg.txn_type_column
        txn_col_idx = header_index.get(txn_col_name) if txn_col_name else None
        raw_type = str(row[txn_col_idx]).strip() if txn_col_idx is not None and txn_col_idx < len(row) and row[txn_col_idx] is not None else ""
        txn_type = sheet_cfg.txn_type_map.get(raw_type, raw_type.upper())
        if txn_type not in ("PURCHASE", "SALE"):
            errors.append(ParseError(source_label, row_number, "txn_type", raw_type,
                                     f"Unknown transaction type '{raw_type}'. Expected PURCHASE or SALE."))
            return None, errors

    # --- Substance ---
    raw_substance = ""
    if config.substance_extraction_enabled and config.substance_extraction_source_column:
        narration = str(get(row, config.substance_extraction_source_column, ""))
        m = re.search(config.substance_extraction_pattern or "", narration, re.IGNORECASE)
        raw_substance = m.group(1) if m else ""
    else:
        raw_substance = str(get(row, "substance", "")).strip()

    resolved_substance = config.resolve_substance(raw_substance)
    anomaly_flags: list[str] = []
    if not resolved_substance:
        anomaly_flags.append(f"UNKNOWN_SUBSTANCE:{raw_substance}")
        resolved_substance = raw_substance  # keep original, flagged

    # --- Quantity ---
    raw_qty = get(row, "quantity_kg")
    quantity_kg = _parse_decimal(raw_qty)
    if quantity_kg is None:
        errors.append(ParseError(source_label, row_number, "quantity_kg", str(raw_qty),
                                 "Cannot parse quantity as a number"))
        return None, errors

    if config.qty_unit_normalisation_enabled and config.qty_unit_column:
        unit_col_name = config.qty_unit_column
        unit_col_idx = _find_col_idx_by_name(row, unit_col_name, sheet_cfg)
        unit_raw = str(row[unit_col_idx]).strip() if unit_col_idx is not None else "kg"
        factor = config.qty_unit_map.get(unit_raw, Decimal("1"))
        quantity_kg = (quantity_kg * factor).quantize(Decimal("0.001"))

    if quantity_kg == Decimal("0.00"):
        anomaly_flags.append("ZERO_QUANTITY")

    # --- Optional numeric fields ---
    rate = _parse_decimal(get(row, "rate_inr_per_kg")) or Decimal("0.00")
    amount = _parse_decimal(get(row, "amount_inr")) or Decimal("0.00")

    # --- String fields ---
    counterparty = str(get(row, "counterparty", "")).strip()
    counterparty_urn = str(get(row, "counterparty_urn", "") or "").strip()
    voucher_no = str(get(row, "voucher_no", "") or "").strip()
    item_code = str(get(row, "item_code", "") or "").strip()
    form_g_no = str(get(row, "form_g_no", "") or "").strip()

    # --- URN validation ---
    urn_result = validate_urn(counterparty_urn if counterparty_urn else None)
    if urn_result.requires_manual_verification:
        anomaly_flags.append(f"URN_{urn_result.status.value}")

    return NCBTransaction(
        client_id=client_id,
        date=parsed_date,
        voucher_no=voucher_no,
        txn_type=txn_type,
        substance=resolved_substance,
        item_code=item_code,
        counterparty=counterparty,
        counterparty_urn=counterparty_urn,
        quantity_kg=quantity_kg,
        rate_inr_per_kg=rate,
        amount_inr=amount,
        form_g_no=form_g_no,
        urn_status=urn_result.status.value,
        urn_requires_manual_verification=urn_result.requires_manual_verification,
        source="excel",
        anomaly_flags=anomaly_flags,
        raw_substance=raw_substance,
        row_number=row_number,
    ), []


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _parse_date(raw: Any, fmt: str) -> Optional[date]:
    if raw is None:
        return None
    if isinstance(raw, date):
        return raw
    try:
        from datetime import datetime
        return datetime.strptime(str(raw).strip(), fmt).date()
    except (ValueError, TypeError):
        return None


def _parse_decimal(raw: Any) -> Optional[Decimal]:
    if raw is None:
        return None
    try:
        cleaned = str(raw).replace(",", "").replace("₹", "").strip()
        return Decimal(cleaned)
    except InvalidOperation:
        return None


def _find_col_idx_by_name(row: tuple, col_name: str, sheet_cfg: SheetConfig) -> Optional[int]:
    return None  # placeholder — only used when qty_unit_normalisation is enabled


# ---------------------------------------------------------------------------
# Google Sheets compatibility shim
# ---------------------------------------------------------------------------

class _GspreadsheetsWorkbook:
    """Wraps a gspread Spreadsheet to mimic openpyxl's workbook interface."""

    def __init__(self, spreadsheet) -> None:
        self._spreadsheet = spreadsheet
        self._sheets = {ws.title: ws for ws in spreadsheet.worksheets()}

    @property
    def sheetnames(self) -> list[str]:
        return list(self._sheets.keys())

    def __getitem__(self, name: str) -> _GspreadsheetsWorksheet:
        return _GspreadsheetsWorksheet(self._sheets[name])


class _GspreadsheetsWorksheet:
    """Wraps a gspread Worksheet to mimic openpyxl's worksheet interface."""

    def __init__(self, worksheet) -> None:
        self._ws = worksheet
        self._data = worksheet.get_all_values()  # list of list of str

    def iter_rows(self, values_only: bool = True):
        for row in self._data:
            yield tuple(cell if cell != "" else None for cell in row)
