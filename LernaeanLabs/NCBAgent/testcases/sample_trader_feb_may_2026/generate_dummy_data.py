"""
Generates 300 synthetic ledger entries for a fictional Schedule A chemical trader
covering Feb 10 – May 9, 2026 (90 days).

Outputs:
  dummy_ledger.xlsx    — Excel ledger (human-readable, mirrors what a consultant would maintain)
  tally_export.xml     — Tally Day Book XML covering the same transactions

Trader profile:
  Company : Mehta Chemical Traders Pvt Ltd
  Location: Ankleshwar GIDC, Bharuch, Gujarat
  URN     : NCB-GJ-2020-007890
  Handles : Anthranilic Acid, N-Acetylanthranilic Acid, Acetic Anhydride (Schedule A)
            Ephedrine, Pseudoephedrine (Schedule A — pharma grade, smaller volumes)

Deliberate anomalies planted (for testing the parser + register module):
  [A1] Entry #87  — counterparty URN has invalid format (5 digits instead of 6)
  [A2] Entry #143 — counterparty URN is blank (missing)
  [A3] Entry #201 — quantity rounded to zero after conversion (edge case)
  [A4] Mar 15 gap — no entries on this date (nil-transaction day must be auto-filled)
"""

import random
import sys
from dataclasses import dataclass
from datetime import date, timedelta
from decimal import ROUND_HALF_UP, Decimal
from pathlib import Path
from xml.dom import minidom
from xml.etree import ElementTree as ET

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    sys.exit("Install openpyxl first: pip install openpyxl")

random.seed(42)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

CLIENT_NAME = "Mehta Chemical Traders Pvt Ltd"
CLIENT_URN = "NCB-GJ-2020-007890"
CLIENT_GSTIN = "24AABCM1234F1ZK"

START_DATE = date(2026, 2, 10)
END_DATE = date(2026, 5, 9)
TOTAL_ENTRIES = 300

ANOMALY_INVALID_URN_IDX = 86   # 0-based → entry #87
ANOMALY_MISSING_URN_IDX = 142  # 0-based → entry #143
ANOMALY_ZERO_QTY_IDX = 200     # 0-based → entry #201
ANOMALY_NIL_DATE = date(2026, 3, 15)   # no entries on this date


@dataclass
class Substance:
    name: str
    item_code: str
    schedule: str
    min_price: Decimal   # INR/kg
    max_price: Decimal
    purchase_lot_min: Decimal   # kg
    purchase_lot_max: Decimal
    sale_lot_min: Decimal
    sale_lot_max: Decimal
    weight: int          # relative sampling weight


SUBSTANCES = [
    Substance(
        name="Anthranilic Acid",
        item_code="ANTHRAC-001",
        schedule="A",
        min_price=Decimal("420"), max_price=Decimal("580"),
        purchase_lot_min=Decimal("100"), purchase_lot_max=Decimal("500"),
        sale_lot_min=Decimal("25"), sale_lot_max=Decimal("200"),
        weight=40,
    ),
    Substance(
        name="N-Acetylanthranilic Acid",
        item_code="N-ANTHRAC-001",
        schedule="A",
        min_price=Decimal("680"), max_price=Decimal("950"),
        purchase_lot_min=Decimal("50"), purchase_lot_max=Decimal("300"),
        sale_lot_min=Decimal("20"), sale_lot_max=Decimal("100"),
        weight=30,
    ),
    Substance(
        name="Acetic Anhydride",
        item_code="ACETIC-001",
        schedule="A",
        min_price=Decimal("75"), max_price=Decimal("95"),
        purchase_lot_min=Decimal("200"), purchase_lot_max=Decimal("1000"),
        sale_lot_min=Decimal("50"), sale_lot_max=Decimal("500"),
        weight=20,
    ),
    Substance(
        name="Ephedrine",
        item_code="EPHED-001",
        schedule="A",
        min_price=Decimal("3200"), max_price=Decimal("4800"),
        purchase_lot_min=Decimal("5"), purchase_lot_max=Decimal("30"),
        sale_lot_min=Decimal("2"), sale_lot_max=Decimal("15"),
        weight=5,
    ),
    Substance(
        name="Pseudoephedrine",
        item_code="PSEUDO-001",
        schedule="A",
        min_price=Decimal("4500"), max_price=Decimal("6500"),
        purchase_lot_min=Decimal("3"), purchase_lot_max=Decimal("20"),
        sale_lot_min=Decimal("1"), sale_lot_max=Decimal("10"),
        weight=5,
    ),
]

SUBSTANCE_WEIGHTS = [s.weight for s in SUBSTANCES]


@dataclass
class Counterparty:
    name: str
    urn: str
    gstin: str
    txn_type: str   # "SUPPLIER" | "BUYER" | "BOTH"


COUNTERPARTIES = [
    Counterparty("Jackson Chemical Industries", "NCB-GJ-2019-001234", "24AACJ5678K1Z3", "BOTH"),
    Counterparty("Nirav Dyes Pvt Ltd", "NCB-GJ-2020-002567", "24AADND9012L1Z8", "BUYER"),
    Counterparty("Ishita Industries", "NCB-GJ-2018-003891", "24AAEII3456M1Z5", "BUYER"),
    Counterparty("GNFC Ltd", "NCB-GJ-2015-004521", "24AABCG7890N1Z2", "SUPPLIER"),
    Counterparty("Link Pharma Chem Ltd", "NCB-GJ-2021-005678", "24AAFPL2345O1Z7", "BUYER"),
    Counterparty("Shree Chemopharma", "NCB-GJ-2020-006234", "24AAGSC6789P1Z4", "BOTH"),
    Counterparty("Himalaya Chemicals", "NCB-GJ-2019-008123", "24AAHCH1234Q1Z9", "BUYER"),
    Counterparty("Vitrag Chemicals", "NCB-GJ-2022-009456", "24AAIVC5678R1Z6", "BUYER"),
    Counterparty("Aarti Industries Ltd", "NCB-GJ-2012-011890", "24AABCA9012S1Z1", "SUPPLIER"),
    Counterparty("Deepak Nitrite Ltd", "NCB-GJ-2014-012345", "24AABCD3456T1Z8", "SUPPLIER"),
    Counterparty("Jubilant Life Sciences", "NCB-UP-2013-021456", "09AABCJ7890U1Z5", "SUPPLIER"),
    Counterparty("Dev International", "NCB-GJ-2017-031789", "24AAKDI2345V1Z2", "BOTH"),
    Counterparty("Endemic India Chemicals", "NCB-GJ-2021-041023", "24AALEI6789W1Z9", "BUYER"),
    Counterparty("Kanoria Chemicals", "NCB-GJ-2016-051234", "24AAMKC1234X1Z6", "SUPPLIER"),
    Counterparty("Laxmi Organic Industries", "NCB-MH-2018-061567", "27AABLO5678Y1Z3", "SUPPLIER"),
]

SUPPLIERS = [c for c in COUNTERPARTIES if c.txn_type in ("SUPPLIER", "BOTH")]
BUYERS = [c for c in COUNTERPARTIES if c.txn_type in ("BUYER", "BOTH")]


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def random_decimal(lo: Decimal, hi: Decimal, places: int = 2) -> Decimal:
    val = lo + Decimal(str(random.random())) * (hi - lo)
    quant = Decimal("0.01") if places == 2 else Decimal("1")
    return val.quantize(quant, rounding=ROUND_HALF_UP)


def random_qty(lo: Decimal, hi: Decimal) -> Decimal:
    # Round to nearest 0.5 kg for realism
    raw = random_decimal(lo, hi, places=1)
    return (raw * 2).quantize(Decimal("1"), rounding=ROUND_HALF_UP) / 2


def form_g_number(date_: date, seq: int) -> str:
    return f"FG/{date_.strftime('%Y%m')}/{seq:04d}"


def voucher_number(txn_type: str, counter: int) -> str:
    prefix = "PUR" if txn_type == "PURCHASE" else "SAL"
    return f"{prefix}/2026/{counter:04d}"


def all_dates() -> list[date]:
    days = []
    d = START_DATE
    while d <= END_DATE:
        if d != ANOMALY_NIL_DATE:
            days.append(d)
        d += timedelta(days=1)
    return days


# ---------------------------------------------------------------------------
# Entry generation
# ---------------------------------------------------------------------------

@dataclass
class LedgerEntry:
    entry_no: int
    date: date
    voucher_no: str
    txn_type: str          # PURCHASE | SALE
    substance: str
    item_code: str
    counterparty: str
    counterparty_urn: str
    quantity_kg: Decimal
    rate_inr_per_kg: Decimal
    amount_inr: Decimal
    form_g_no: str
    urn_valid: str         # YES | NO (format check) | MISSING
    anomaly_flag: str      # blank or anomaly code


def generate_entries() -> list[LedgerEntry]:
    available_dates = all_dates()
    entries: list[LedgerEntry] = []

    # Spread 300 entries across available dates with mild clustering
    date_pool: list[date] = []
    for d in available_dates:
        # 1–5 entries per day, weighted slightly toward middle of week
        count = random.choices([1, 2, 3, 4, 5], weights=[5, 20, 35, 25, 15])[0]
        date_pool.extend([d] * count)

    random.shuffle(date_pool)
    selected_dates = sorted(date_pool[:TOTAL_ENTRIES])

    pur_counter = 1
    sal_counter = 1
    fg_counter = 1

    for i, entry_date in enumerate(selected_dates):
        substance = random.choices(SUBSTANCES, weights=SUBSTANCE_WEIGHTS)[0]
        txn_type = random.choice(["PURCHASE", "PURCHASE", "SALE"])  # slight purchase bias

        if txn_type == "PURCHASE":
            party = random.choice(SUPPLIERS)
            qty = random_qty(substance.purchase_lot_min, substance.purchase_lot_max)
            vno = voucher_number("PURCHASE", pur_counter)
            pur_counter += 1
        else:
            party = random.choice(BUYERS)
            qty = random_qty(substance.sale_lot_min, substance.sale_lot_max)
            vno = voucher_number("SALE", sal_counter)
            sal_counter += 1

        rate = random_decimal(substance.min_price, substance.max_price)
        amount = (qty * rate).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        urn = party.urn
        urn_valid = "YES"
        anomaly = ""

        # Plant anomalies
        if i == ANOMALY_INVALID_URN_IDX:
            urn = "NCB-GJ-2019-01234"   # 5 digits — missing one digit [A1]
            urn_valid = "NO"
            anomaly = "A1:INVALID_URN_FORMAT"

        elif i == ANOMALY_MISSING_URN_IDX:
            urn = ""                     # blank URN [A2]
            urn_valid = "MISSING"
            anomaly = "A2:MISSING_URN"

        elif i == ANOMALY_ZERO_QTY_IDX:
            qty = Decimal("0.00")        # zero qty edge case [A3]
            amount = Decimal("0.00")
            anomaly = "A3:ZERO_QUANTITY"

        entries.append(LedgerEntry(
            entry_no=i + 1,
            date=entry_date,
            voucher_no=vno,
            txn_type=txn_type,
            substance=substance.name,
            item_code=substance.item_code,
            counterparty=party.name,
            counterparty_urn=urn,
            quantity_kg=qty,
            rate_inr_per_kg=rate,
            amount_inr=amount,
            form_g_no=form_g_number(entry_date, fg_counter),
            urn_valid=urn_valid,
            anomaly_flag=anomaly,
        ))
        fg_counter += 1

    return entries


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
ANOMALY_FILL = PatternFill("solid", fgColor="FFD700")
ALT_FILL = PatternFill("solid", fgColor="EBF1F8")

COLUMNS = [
    ("Entry No",         10),
    ("Date",             12),
    ("Voucher No",       16),
    ("Txn Type",         12),
    ("Substance",        28),
    ("Item Code",        18),
    ("Counterparty",     32),
    ("Counterparty URN", 22),
    ("Qty (kg)",         12),
    ("Rate (₹/kg)",      12),
    ("Amount (₹)",       14),
    ("Form-G No",        18),
    ("URN Valid",        12),
    ("Anomaly Flag",     22),
]


def write_excel(entries: list[LedgerEntry], output_path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NCB Ledger"

    # Header row
    for col_idx, (header, width) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[cell.column_letter].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    for row_idx, e in enumerate(entries, start=2):
        row_data = [
            e.entry_no,
            e.date.strftime("%d-%b-%Y"),
            e.voucher_no,
            e.txn_type,
            e.substance,
            e.item_code,
            e.counterparty,
            e.counterparty_urn,
            float(e.quantity_kg),
            float(e.rate_inr_per_kg),
            float(e.amount_inr),
            e.form_g_no,
            e.urn_valid,
            e.anomaly_flag,
        ]
        fill = ANOMALY_FILL if e.anomaly_flag else (ALT_FILL if row_idx % 2 == 0 else None)
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal="left")

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Client"
    ws2["B1"] = CLIENT_NAME
    ws2["A2"] = "URN"
    ws2["B2"] = CLIENT_URN
    ws2["A3"] = "Period"
    ws2["B3"] = f"{START_DATE.strftime('%d %b %Y')} – {END_DATE.strftime('%d %b %Y')}"
    ws2["A4"] = "Total Entries"
    ws2["B4"] = len(entries)
    ws2["A5"] = "Nil-Transaction Date"
    ws2["B5"] = ANOMALY_NIL_DATE.strftime("%d %b %Y")
    ws2["A6"] = "Anomalies Planted"
    ws2["B6"] = 3
    ws2["A8"] = "Anomaly"
    ws2["B8"] = "Entry No"
    ws2["C8"] = "Description"
    ws2["A9"] = "A1"
    ws2["B9"] = ANOMALY_INVALID_URN_IDX + 1
    ws2["C9"] = "Invalid URN format (5 digits in sequence — should be 6)"
    ws2["A10"] = "A2"
    ws2["B10"] = ANOMALY_MISSING_URN_IDX + 1
    ws2["C10"] = "URN field is blank"
    ws2["A11"] = "A3"
    ws2["B11"] = ANOMALY_ZERO_QTY_IDX + 1
    ws2["C11"] = "Zero quantity — edge case for parser"

    for col in ["A", "B", "C"]:
        ws2.column_dimensions[col].width = 30

    wb.save(output_path)
    print(f"[Excel] Written: {output_path}  ({len(entries)} rows)")


# ---------------------------------------------------------------------------
# Tally XML output
# ---------------------------------------------------------------------------

def _add_ledger_entry(voucher_el: ET.Element, ledger_name: str, amount: Decimal, is_party: bool) -> None:
    el = ET.SubElement(voucher_el, "ALLLEDGERENTRIES.LIST")
    ET.SubElement(el, "LEDGERNAME").text = ledger_name
    ET.SubElement(el, "ISDEEMEDPOSITIVE").text = "No"
    ET.SubElement(el, "ISLASTDEEMEDPOSITIVE").text = "No"
    ET.SubElement(el, "ISPARTYLEDGER").text = "Yes" if is_party else "No"
    ET.SubElement(el, "AMOUNT").text = str(amount)


def _add_inventory_entry(
    voucher_el: ET.Element,
    item_name: str,
    qty: Decimal,
    rate: Decimal,
    amount: Decimal,
) -> None:
    el = ET.SubElement(voucher_el, "ALLINVENTORYENTRIES.LIST")
    ET.SubElement(el, "STOCKITEMNAME").text = item_name
    ET.SubElement(el, "ISDEEMEDPOSITIVE").text = "No"
    ET.SubElement(el, "RATE").text = f"{rate}/Kg"
    ET.SubElement(el, "AMOUNT").text = str(amount)
    ET.SubElement(el, "ACTUALQTY").text = f"{qty} Kg"
    ET.SubElement(el, "BILLEDQTY").text = f"{qty} Kg"


def write_tally_xml(entries: list[LedgerEntry], output_path: Path) -> None:
    envelope = ET.Element("ENVELOPE")

    header = ET.SubElement(envelope, "HEADER")
    ET.SubElement(header, "TALLYREQUEST").text = "Export Data"

    body = ET.SubElement(envelope, "BODY")
    export_data = ET.SubElement(body, "EXPORTDATA")

    req_desc = ET.SubElement(export_data, "REQUESTDESC")
    ET.SubElement(req_desc, "REPORTNAME").text = "Day Book"
    static_vars = ET.SubElement(req_desc, "STATICVARIABLES")
    ET.SubElement(static_vars, "SVFROMDATE").text = START_DATE.strftime("%Y%m%d")
    ET.SubElement(static_vars, "SVTODATE").text = END_DATE.strftime("%Y%m%d")
    ET.SubElement(static_vars, "SVEXPORTFORMAT").text = "$$SysName:XML"
    ET.SubElement(static_vars, "SVCURRENTCOMPANY").text = CLIENT_NAME

    req_data = ET.SubElement(export_data, "REQUESTDATA")

    for e in entries:
        msg = ET.SubElement(req_data, "TALLYMESSAGE")
        msg.set("xmlns:UDF", "TallyUDF")

        vch_type = "Purchase" if e.txn_type == "PURCHASE" else "Sales"
        voucher = ET.SubElement(msg, "VOUCHER")
        voucher.set("VCHTYPE", vch_type)
        voucher.set("ACTION", "Create")
        voucher.set("OBJVIEW", "Invoice Voucher View")

        ET.SubElement(voucher, "DATE").text = e.date.strftime("%Y%m%d")
        ET.SubElement(voucher, "GUID").text = f"NCB-{e.entry_no:06d}"
        ET.SubElement(voucher, "VOUCHERTYPENAME").text = vch_type
        ET.SubElement(voucher, "VOUCHERNUMBER").text = e.voucher_no
        ET.SubElement(voucher, "PARTYLEDGERNAME").text = e.counterparty
        ET.SubElement(voucher, "NARRATION").text = (
            f"{e.txn_type.title()} of {e.substance} | "
            f"Form-G: {e.form_g_no} | "
            f"Counterparty URN: {e.counterparty_urn or 'NOT PROVIDED'}"
        )

        # UDF fields for NCB-specific data (custom fields in Tally)
        udf = ET.SubElement(voucher, "UDF:COUNTERPARTYURN.LIST")
        ET.SubElement(udf, "UDF:COUNTERPARTYURN").text = e.counterparty_urn
        udf2 = ET.SubElement(voucher, "UDF:FORMGNO.LIST")
        ET.SubElement(udf2, "UDF:FORMGNO").text = e.form_g_no

        # Accounting entry: debit purchase / credit sales account
        # Tally convention: purchases are negative, sales are positive from ledger POV
        ledger_amount = -e.amount_inr if e.txn_type == "PURCHASE" else e.amount_inr
        _add_ledger_entry(voucher, f"{e.substance} {'Purchases' if e.txn_type == 'PURCHASE' else 'Sales'}", ledger_amount, is_party=False)
        # Party leg (supplier credit / buyer debit)
        party_amount = e.amount_inr if e.txn_type == "PURCHASE" else -e.amount_inr
        _add_ledger_entry(voucher, e.counterparty, party_amount, is_party=True)

        # Inventory entry
        inv_amount = -e.amount_inr if e.txn_type == "PURCHASE" else e.amount_inr
        _add_inventory_entry(voucher, e.substance, e.quantity_kg, e.rate_inr_per_kg, inv_amount)

    # Pretty-print via minidom
    raw = ET.tostring(envelope, encoding="unicode")
    pretty = minidom.parseString(raw).toprettyxml(indent="  ", encoding="UTF-8")
    output_path.write_bytes(pretty)
    print(f"[XML]   Written: {output_path}  ({len(entries)} vouchers)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    out_dir = Path(__file__).parent

    entries = generate_entries()

    # Verify anomaly dates and indexing
    assert entries[ANOMALY_INVALID_URN_IDX].anomaly_flag.startswith("A1"), "A1 anomaly not planted correctly"
    assert entries[ANOMALY_MISSING_URN_IDX].anomaly_flag.startswith("A2"), "A2 anomaly not planted correctly"
    assert entries[ANOMALY_ZERO_QTY_IDX].anomaly_flag.startswith("A3"), "A3 anomaly not planted correctly"

    # Check nil-transaction date is absent
    entry_dates = {e.date for e in entries}
    assert ANOMALY_NIL_DATE not in entry_dates, f"{ANOMALY_NIL_DATE} should have no entries"

    write_excel(entries, out_dir / "dummy_ledger.xlsx")
    write_tally_xml(entries, out_dir / "tally_export.xml")

    # Quick stats
    purchases = [e for e in entries if e.txn_type == "PURCHASE"]
    sales = [e for e in entries if e.txn_type == "SALE"]
    total_purchase_value = sum(e.amount_inr for e in purchases)
    total_sale_value = sum(e.amount_inr for e in sales)

    print(f"\n--- Summary ---")
    print(f"Total entries  : {len(entries)}")
    print(f"  Purchases    : {len(purchases)}  (₹{total_purchase_value:,.2f})")
    print(f"  Sales        : {len(sales)}  (₹{total_sale_value:,.2f})")
    print(f"Date range     : {entries[0].date} → {entries[-1].date}")
    print(f"Nil-txn date   : {ANOMALY_NIL_DATE}  (no entries — register module must auto-fill)")
    print(f"Anomalies      : A1=entry#{ANOMALY_INVALID_URN_IDX+1}, A2=entry#{ANOMALY_MISSING_URN_IDX+1}, A3=entry#{ANOMALY_ZERO_QTY_IDX+1}")
    substance_counts: dict[str, int] = {}
    for e in entries:
        substance_counts[e.substance] = substance_counts.get(e.substance, 0) + 1
    print(f"\nBy substance:")
    for sub, cnt in sorted(substance_counts.items(), key=lambda x: -x[1]):
        print(f"  {sub:<35} {cnt:>3} entries")
